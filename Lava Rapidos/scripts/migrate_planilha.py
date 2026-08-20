# -*- coding: utf-8 -*-
"""
Migra os dados da planilha DUH.xlsm para um arquivo seed.sql (UTF-8) pronto
para colar/rodar no SQL Editor do Supabase APÓS aplicar o schema.sql.

Uso:
    pip install openpyxl
    python migrate_planilha.py

Gera: ../supabase/seed.sql  (NÃO versionar — contém dados pessoais / LGPD)
"""
import os, re, uuid, datetime, warnings
import openpyxl

warnings.filterwarnings("ignore")
HERE = os.path.dirname(os.path.abspath(__file__))
XLSM = os.path.join(HERE, "..", "DUH.xlsm")
OUT  = os.path.join(HERE, "..", "supabase", "seed.sql")

wb = openpyxl.load_workbook(XLSM, data_only=True)


def q(v):
    """Formata valor para SQL."""
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return "NULL"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    if isinstance(v, datetime.time):
        return "'" + v.strftime("%H:%M:%S") + "'"
    return "'" + str(v).strip().replace("'", "''") + "'"


def num(v):
    """Converte valor monetário (inclusive '165,52' / 'R$ 1.200,00') para float SQL."""
    if v is None or v == "":
        return "0"
    if isinstance(v, (int, float)):
        return repr(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    # remove separador de milhar e troca vírgula decimal por ponto
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return repr(float(s))
    except ValueError:
        return "0"


def norm(s):
    return re.sub(r"\s+", " ", str(s).strip()).upper() if s else ""


def is_yuri(obs):
    return obs is not None and "YURI" in str(obs).upper()


rows_out = []
rows_out.append("-- seed.sql gerado de DUH.xlsm — contém dados pessoais (LGPD). NÃO versionar.")
rows_out.append("begin;")

# ---------------------------------------------------------------- CLIENTES + CARROS
cliente_id_by_name = {}
ws = wb["Clientes"]
data = list(ws.iter_rows(values_only=True))[1:]
nc = ncar = 0
for r in data:
    _id, nome, tel, veic, placa, origem, obs = (list(r) + [None] * 7)[:7]
    if not nome:
        continue
    cid = str(uuid.uuid4())
    cliente_id_by_name.setdefault(norm(nome), cid)
    base = is_yuri(obs)
    rows_out.append(
        "insert into clientes (id,nome,telefone,origem,base_antiga,observacoes) values "
        f"({q(cid)},{q(nome)},{q(tel)},{q(origem or 'PARTICULAR')},{str(base).lower()},{q(obs)});"
    )
    nc += 1
    if placa:
        rows_out.append(
            "insert into carros (cliente_id,placa,veiculo) values "
            f"({q(cid)},{q(placa)},{q(veic)});"
        )
        ncar += 1

# ---------------------------------------------------------------- PARCEIROS
parceiro_id_by_name = {}
ws = wb["Parceiros"]
for r in list(ws.iter_rows(values_only=True))[3:]:
    nome = r[0]
    if not nome or norm(nome) in ("RÓTULOS DE LINHA", "TOTAL GERAL"):
        continue
    pid = str(uuid.uuid4())
    parceiro_id_by_name[norm(nome)] = pid
    rows_out.append(f"insert into parceiros (id,nome) values ({q(pid)},{q(nome)});")
np = len(parceiro_id_by_name)

# ---------------------------------------------------------------- FUNCIONÁRIOS
func_id_by_name = {}
nomes_func = set()
for r in list(wb["Presença"].iter_rows(values_only=True))[2:]:
    if r[1]:
        nomes_func.add(norm(r[1]))
for r in list(wb["Histórico Colaboradores"].iter_rows(values_only=True))[2:]:
    if r[2]:
        nomes_func.add(norm(r[2]))
nomes_func.discard("(VAZIO)")
nomes_func.discard("COLABORADOR")
for nome in sorted(nomes_func):
    fid = str(uuid.uuid4())
    func_id_by_name[nome] = fid
    rows_out.append(f"insert into funcionarios (id,nome) values ({q(fid)},{q(nome.title())});")
nf = len(func_id_by_name)

# ---------------------------------------------------------------- ATENDIMENTOS
ws = wb["Atendimentos"]
all_rows = list(ws.iter_rows(values_only=True))
# header em all_rows[12]; dados a partir de 13
atend_id_by_os = {}
na = 0
extra_cli = extra_par = 0
for r in all_rows[13:]:
    r = (list(r) + [None] * 14)[:14]
    data_, os_, cliente, tel, veic, placa, origem, serv, valor, fp, stpg, datapg, obs, mes = r
    if not os_:
        continue
    tipo = "PARCEIRO" if norm(origem) == "PARCEIRO" else "PARTICULAR"
    cid = pid = "NULL"
    if tipo == "PARTICULAR":
        key = norm(cliente)
        if key and key not in cliente_id_by_name:
            new = str(uuid.uuid4())
            cliente_id_by_name[key] = new
            rows_out.append(
                "insert into clientes (id,nome,telefone,base_antiga) values "
                f"({q(new)},{q(cliente)},{q(tel)},{str(is_yuri(obs)).lower()});"
            )
            extra_cli += 1
        cid = q(cliente_id_by_name.get(key)) if key else "NULL"
    else:
        key = norm(cliente)
        if key and key not in parceiro_id_by_name:
            new = str(uuid.uuid4())
            parceiro_id_by_name[key] = new
            rows_out.append(f"insert into parceiros (id,nome) values ({q(new)},{q(cliente)});")
            extra_par += 1
        pid = q(parceiro_id_by_name.get(key)) if key else "NULL"

    aid = str(uuid.uuid4())
    atend_id_by_os[norm(os_)] = (aid, is_yuri(obs))
    status = "PAGO" if norm(stpg) == "PAGO" else "PENDENTE"
    rows_out.append(
        "insert into atendimentos (id,os_numero,data,tipo,cliente_id,parceiro_id,veiculo,placa,"
        "servicos,valor,forma_pgto,status_pg,data_pg,base_antiga,observacoes) values ("
        f"{q(aid)},{q(os_)},{q(data_)},{q(tipo)},{cid},{pid},{q(veic)},{q(placa)},{q(serv)},"
        f"{num(valor)},{q(fp)},{q(status)},{q(datapg)},{str(is_yuri(obs)).lower()},{q(obs)});"
    )
    na += 1

# ---------------------------------------------------------------- FINANCEIRO
ws = wb["Financeiro"]
nfin = 0
for r in list(ws.iter_rows(values_only=True))[2:]:
    r = (list(r) + [None] * 10)[:10]
    data_, tipo, os_, cliente, desc, valor, fp, valreal, obs, mes = r
    if not tipo:
        continue
    t = "ENTRADA" if "ENTRADA" in norm(tipo) else "SAIDA"
    aid_sql = "NULL"
    base = is_yuri(obs)
    if os_ and norm(os_) in atend_id_by_os:
        aid, abase = atend_id_by_os[norm(os_)]
        aid_sql = q(aid)
        base = base or abase
    rows_out.append(
        "insert into financeiro (data,tipo,atendimento_id,descricao,valor,forma_pgto,base_antiga,observacoes) "
        f"values ({q(data_)},{q(t)},{aid_sql},{q(desc)},{num(valor)},{q(fp)},{str(base).lower()},{q(obs)});"
    )
    nfin += 1

# ---------------------------------------------------------------- PRESENÇA
ws = wb["Presença"]
npres = 0
seen = set()
for r in list(ws.iter_rows(values_only=True))[2:]:
    r = (list(r) + [None] * 8)[:8]
    data_, colab, status, hora = r[0], r[1], r[2], r[3]
    if not (data_ and colab and status):
        continue
    fid = func_id_by_name.get(norm(colab))
    if not fid:
        continue
    st = "PRESENTE" if "PRESENTE" in norm(status) else ("FALTA" if "FALTA" in norm(status) else None)
    if not st:
        continue
    d = data_.strftime("%Y-%m-%d") if hasattr(data_, "strftime") else str(data_)
    if (d, fid) in seen:
        continue
    seen.add((d, fid))
    rows_out.append(
        "insert into presenca (data,funcionario_id,status,hora) values "
        f"({q(data_)},{q(fid)},{q(st)},{q(hora)}) on conflict (data,funcionario_id) do nothing;"
    )
    npres += 1

# ---------------------------------------------------------------- MOVIMENTOS COLABORADOR
ws = wb["Histórico Colaboradores"]
nmov = 0
for r in list(ws.iter_rows(values_only=True))[2:]:
    r = (list(r) + [None] * 6)[:6]
    data_, mes, colab, tipo, valor, obs = r
    if not (colab and tipo):
        continue
    t = "Pagamento" if "PAGAMENTO" in norm(tipo) else "Vale"
    fid = func_id_by_name.get(norm(colab))
    fid_sql = q(fid) if fid else "NULL"
    rows_out.append(
        "insert into colaborador_movimentos (data,funcionario_id,funcionario_nome,tipo,valor,observacao) "
        f"values ({q(data_)},{fid_sql},{q(colab)},{q(t)},{num(valor)},{q(obs)});"
    )
    nmov += 1

rows_out.append("commit;")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(rows_out) + "\n")

print(f"OK -> {OUT}")
print(f"  clientes:        {nc} (+{extra_cli} extras de atendimentos)")
print(f"  carros:          {ncar}")
print(f"  parceiros:       {np} (+{extra_par} extras)")
print(f"  funcionarios:    {nf}")
print(f"  atendimentos:    {na}")
print(f"  financeiro:      {nfin}")
print(f"  presenca:        {npres}")
print(f"  mov. colab:      {nmov}")
